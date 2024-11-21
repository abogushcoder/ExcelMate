import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def read_txt_file(filename: str):
    try:
        big_lst = []
        with open(filename, 'r') as file:
            for line in file:
                big_lst.append([line.strip()])
            return big_lst
    except Exception as e:
        print(f"Error: {e}")


def get_info(big_lst: list):
    char_in_col = 1
    col_num = 1
    col_to_char_dict = {}
    while char_in_col != 0:
        char_in_col = int(input(f"Input how many characters are in column {col_num} (if done enter 0): "))
        if char_in_col == 0:
            break
        col_name = input(f"Input the column name for column {col_num}: ")
        format = input(f"What format do you want for column {col_num}('c' for currency, 'd' for date, 't' for time, 'g' for general): ")
        col_to_char_dict[col_num] = (char_in_col, col_name, format)
        for lst in big_lst:
            part1 = lst[-1][:char_in_col]
            part2 = lst[-1][char_in_col:]
            lst.pop(-1)
            lst.append(part1)
            if len(part2) > 0:
                lst.append(part2)
        col_num += 1
    return col_to_char_dict


def get_row_of_col_names(col_to_char_dict: dict):
    col_name_lst = []
    for i in range(1, len(col_to_char_dict) + 1):
        col_name = col_to_char_dict[i][1]
        col_name_lst.append(col_name)
    return col_name_lst


def convert_currency_cols_to_float(col_to_char_dict: dict, big_lst: list):
    currency_index_lst = []
    for col_num in col_to_char_dict:
        if 'c' in col_to_char_dict[col_num]:
            currency_index_lst.append(col_num - 1)
    for lst in big_lst:
        for index in currency_index_lst:
            lst[index] = float(lst[index])


def create_workbook(big_lst: list, col_name_lst: list, col_to_char_dict: dict, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.append(col_name_lst)
    for lst in big_lst:
        ws.append(lst)

    number_of_columns = len(col_to_char_dict)
    for i in range(1, number_of_columns + 1):
        col_reference = get_column_letter(i)
        for cell in ws[col_reference]:
            if col_to_char_dict[i][2] == 'c':
                cell.number_format = '$#,##0.00'  # Currency format
            elif col_to_char_dict[i][2] == 'd':
                cell.number_format = 'mm/dd/yyyy'  # Date format
            elif col_to_char_dict[i][2] == 't':
                cell.number_format = 'hh:mm:ss'  # Time format

    filename = filename[:-4] + '.xlsx'
    wb.save(filename)


def save_input_to_file(col_to_char_dict: dict):
    y_n = input("Do you want to save your input (y/n): ")
    if y_n.lower() == 'y':
        filename = input("Input the name of the file that you want to save to: ")
        with open(filename, 'w') as file:
            json.dump(col_to_char_dict, file)


def load_input_from_file(filename: str):
    try:
        with open(filename, 'r') as file:
            col_to_char_dict = json.load(file)
            col_to_char_dict = {int(k): tuple(v) for k, v in col_to_char_dict.items()}  # Convert keys to int and values to tuples
            return col_to_char_dict
    except Exception as e:
        print(f"Error: {e}")
        return None


def split_big_lst_if_from_save(col_to_char_dict: dict, big_lst: list):
    for column in col_to_char_dict:
        char_in_col = col_to_char_dict[column][0]
        for lst in big_lst:
            part1 = lst[-1][:char_in_col]
            part2 = lst[-1][char_in_col:]
            lst.pop(-1)
            lst.append(part1)
            if len(part2) > 0:
                lst.append(part2)


def main():
    print("Make sure that the .txt file that you use is in the same directory as this Python file\n")
    filename = input("Input the name of the .txt file: ")
    big_lst = read_txt_file(filename)
    if big_lst is None:
        print("Failed to read file. Exiting.")
        return
    use_saved_input = input("Do you want to use the saved input from a file? (y/n): ").lower()
    if use_saved_input == 'y':
        input_file_name = input("Input the name of the file that you want to use: ")
        col_to_char_dict = load_input_from_file(input_file_name)
        split_big_lst_if_from_save(col_to_char_dict, big_lst)
        if col_to_char_dict is None:
            print("Failed to load input from file. Exiting.")
            return
    else:
        col_to_char_dict = get_info(big_lst)
        save_input_to_file(col_to_char_dict)

    col_name_lst = get_row_of_col_names(col_to_char_dict)
    convert_currency_cols_to_float(col_to_char_dict, big_lst)
    create_workbook(big_lst, col_name_lst, col_to_char_dict, filename)


main()
